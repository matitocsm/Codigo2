# -*- coding: utf-8 -*-
# pip install pyinstaller requests openpyxl watchdog pandas tqdm
# pyinstaller --onefile --console procesador_contable.py

import os
import re
import time
import calendar
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from tqdm import tqdm
import sys
from tkinter import messagebox, Tk

if len(sys.argv) > 1:
    BASE_DIR = sys.argv[1]
else:
    BASE_DIR = r"C:\\datos\\Practicas"

modo_gui = os.path.exists(os.path.join(BASE_DIR, "__modo_gui__"))

SPANISH_MONTHS = {
    'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
    'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
    'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
}

def parse_fecha(fecha_str: str) -> datetime:
    m = re.search(r'([A-Za-zñÑáéíóúÁÉÍÓÚ]+)\s+(\d{4})', fecha_str)
    if not m:
        raise ValueError(f"No pude parsear la fecha: {fecha_str!r}")
    mes_sp, año = m.groups()
    mes_num = SPANISH_MONTHS.get(mes_sp.lower())
    año, mes = int(año), int(mes_num)
    _, ultimo_dia = calendar.monthrange(año, mes)
    return datetime(año, mes, ultimo_dia).date()

def process_file(path: str) -> pd.DataFrame:
    for _ in range(10):
        try:
            df_raw = pd.read_excel(path, header=None, dtype=str, keep_default_na=False, engine='openpyxl')
            break
        except PermissionError:
            time.sleep(1)
    else:
        raise PermissionError(f"Acceso denegado tras 10 intentos: {path}")

    fecha = parse_fecha(df_raw.iat[4, 0])

    is_hdr = df_raw.apply(
        lambda r: r.astype(str).str.lower().str.contains('código cuenta contable').any(),
        axis=1
    )
    if not is_hdr.any():
        raise ValueError("No encontré 'Código cuenta contable'")
    hr = is_hdr.idxmax()

    df_body = df_raw.iloc[hr:].reset_index(drop=True)
    df_body.columns = df_body.iloc[0]
    df_data = df_body.iloc[1:].reset_index(drop=True)

    code_col = next(c for c in df_data.columns if 'código cuenta contable' in c.lower())
    name_col = next(c for c in df_data.columns if 'nombre' in c.lower() and 'cuenta' in c.lower())
    trans_col = next(c for c in df_data.columns if 'transaccional' in c.lower())

    metrics = ['Saldo inicial', 'Movimiento débito', 'Movimiento crédito', 'Saldo final']
    for m in metrics:
        if m in df_data.columns:
            df_data[m] = pd.to_numeric(df_data[m], errors='coerce').fillna(0)

    name_map = dict(zip(df_data[code_col], df_data[name_col]))

    df_trans = df_data[df_data[trans_col].str.strip().str.lower() == 'sí'].copy()

    df_trans['Clase']     = df_trans[code_col].str[:1]
    df_trans['Grupo']     = df_trans[code_col].str[:2]
    df_trans['Cuenta']    = df_trans[code_col].str[:4]
    df_trans['Subcuenta'] = df_trans[code_col].str[:6]

    df_trans['Auxiliar'] = df_trans[code_col].apply(
        lambda x: x[:8] if len(x) >= 8 else 'no aplica'
    )
    df_trans['Nombre_auxiliar'] = df_trans['Auxiliar'].map(name_map).fillna('no aplica')

    df_trans['Nombre Clase']  = df_trans['Clase'].map(name_map).fillna('no aplica')
    df_trans['Nombre_Grupo']  = df_trans['Grupo'].map(name_map).fillna('no aplica')
    df_trans['Nombre_cuenta'] = df_trans['Cuenta'].map(name_map).fillna('no aplica')
    df_trans['Nombre_sub']    = df_trans['Subcuenta'].map(name_map).fillna('no aplica')

    df_trans['Sucursal']       = df_trans.get('Sucursal', '').replace('', 'no aplica').fillna('no aplica')
    df_trans['Nombre tercero'] = df_trans.get('Nombre tercero', '').replace('', 'no aplica').fillna('no aplica')
    df_trans['Fecha']          = fecha

    df_trans['Categoría'] = df_trans['Clase'].apply(
        lambda c: 'Balance general' if c in {'1', '2', '3', '9'} else 'Estado de Resultado'
    )

    df_trans['Saldo mes'] = df_trans['Movimiento débito'] - df_trans['Movimiento crédito']

    rename_map = {
        'Auxiliar': 'Sub_cuenta',
        'Nombre_auxiliar': 'Nombre_subcuenta',
        'Subcuenta': 'Cuenta',
        'Nombre_sub': 'Nombre_Cuenta',
        'Cuenta': 'Sub_grupo',
        'Nombre_cuenta': 'Nombre_subgrupo'
    }
    df_trans = df_trans.rename(columns=rename_map)

    final_cols = [
        'Categoría', 'Clase', 'Nombre Clase',
        'Grupo', 'Nombre_Grupo',
        'Sub_grupo', 'Nombre_subgrupo',
        'Cuenta', 'Nombre_Cuenta',
        'Sub_cuenta', 'Nombre_subcuenta',
        'Sucursal', 'Nombre tercero',
        'Saldo inicial', 'Movimiento débito',
        'Movimiento crédito', 'Saldo mes', 'Saldo final',
        'Fecha'
    ]
    df_final = df_trans.reindex(columns=final_cols).fillna('no aplica')

    df_final.columns = [
        col.strip().lower().replace(' ', '_').capitalize()
        for col in df_final.columns
    ]

    return df_final

class ExcelHandler(FileSystemEventHandler):
    def __init__(self, watch_dir: str, output_dir: str):
        self.watch_dir  = watch_dir
        self.output_dir = output_dir
        self.final_path = os.path.join(output_dir, "procesado_final.xlsx")

    def on_created(self, event):
        if not event.src_path.lower().endswith(".xlsx"):
            return
        try:
            df_new = process_file(event.src_path)
            fecha_new = str(df_new['Fecha'].iat[0])

            if os.path.exists(self.final_path):
                df_old = pd.read_excel(self.final_path, engine='openpyxl')
                df_old['Fecha'] = pd.to_datetime(df_old['Fecha']).dt.date
                fechas_old = set(df_old['Fecha'].astype(str).unique())

                if fecha_new in fechas_old:
                    if modo_gui:
                        root = Tk()
                        root.withdraw()
                        respuesta = messagebox.askyesno("Reprocesar archivo", f"Ya existe información con la fecha {fecha_new}.\n¿Deseas reprocesar el archivo {os.path.basename(event.src_path)}?")
                        root.destroy()
                    else:
                        respuesta = input(f"¿Deseas REPROCESAR el archivo {os.path.basename(event.src_path)}? (s/n): ").strip().lower() == 's'

                    if not respuesta:
                        print(f"[SKIP] {os.path.basename(event.src_path)} no reprocesado.")
                        return

                    df_old_filtrado = df_old[df_old['Fecha'].astype(str) != fecha_new]
                    df_final = pd.concat([df_old_filtrado, df_new], ignore_index=True)
                else:
                    df_final = pd.concat([df_old, df_new], ignore_index=True)
            else:
                df_final = df_new

            wb = load_workbook(self.final_path) if os.path.exists(self.final_path) else None
            if wb:
                ws = wb.active
                ws.delete_rows(2, ws.max_row)
            else:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.append(list(df_final.columns))

            with tqdm(total=len(df_final), desc="Guardando filas", unit="fila") as pbar:
                fecha_col_idx = df_final.columns.get_loc("Fecha") + 1
                for i, row in enumerate(dataframe_to_rows(df_final, index=False, header=False), start=2):
                    ws.append(row)
                    ws.cell(row=i, column=fecha_col_idx).number_format = numbers.FORMAT_DATE_YYYYMMDD2
                    pbar.update(1)
                wb.save(self.final_path)

            print(f"[OK] Procesado: {os.path.basename(event.src_path)}")

        except PermissionError:
            print(f"[ERROR] Acceso denegado tras 10 intentos: {os.path.basename(event.src_path)}")
        except Exception as e:
            print(f"[ERROR] Procesando {os.path.basename(event.src_path)}: {e}")

if __name__ == "__main__":
    observer = Observer()

    for entry in os.listdir(BASE_DIR):
        sub = os.path.join(BASE_DIR, entry)
        if not os.path.isdir(sub) or entry.lower() == "codigo":
            continue

        out_dir = os.path.join(sub, "salida")
        os.makedirs(out_dir, exist_ok=True)

        handler = ExcelHandler(sub, out_dir)

        for f in os.listdir(sub):
            if f.lower().endswith(".xlsx"):
                evt = type("E", (), {"src_path": os.path.join(sub, f)})
                handler.on_created(evt)

        observer.schedule(handler, sub, recursive=False)

    observer.start()
    print(f"Vigilando subcarpetas en {BASE_DIR} (excepto 'Codigo')…")
    try:
        while True:
            time.sleep(1.0)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()